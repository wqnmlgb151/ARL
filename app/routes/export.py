#coding: utf-8

import json
import re
from collections import Counter
from urllib.parse import quote

from flask import make_response, request
from flask_restx import Resource, Namespace
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.writer.excel import save_virtual_workbook

from app.utils import get_logger, auth
from app.database.repositories import TaskRepository, IPRepository, SiteRepository, DomainRepository

ns = Namespace('export', description="任务报告导出接口")

logger = get_logger()

# 仓库实例（模块级单例，所有导出请求复用）
task_repo = TaskRepository()
ip_repo = IPRepository()
site_repo = SiteRepository()
domain_repo = DomainRepository()

# 支持的导出格式
_EXPORT_FORMATS = {
    'xlsx': None,  # 在模块末尾延迟注册
    'json': None,
}


def register_format(name, writer):
    """注册导出格式（插件扩展点）"""
    _EXPORT_FORMATS[name] = writer


@ns.route('/<string:task_id>')
class ARLExport(Resource):
    @auth
    def get(self, task_id):
        """
        报告导出（支持 ?format=xlsx|json 参数）
        """
        export_format = request.args.get('format', 'xlsx')
        if export_format not in _EXPORT_FORMATS:
            return {"error": f"Unsupported format: {export_format}. Use: {', '.join(_EXPORT_FORMATS)}"}, 400

        task_data = get_task_data(task_id)
        if not task_data:
            return "not found"

        domain = task_data["target"].replace("/", "_")[:20]
        ext = {'xlsx': 'xlsx', 'json': 'json'}.get(export_format, export_format)
        filename = "ARL资产导出报告_{}.{}".format(domain, ext)

        writer = _EXPORT_FORMATS[export_format]
        data, content_type = writer(task_data, task_id)

        response = make_response(data)
        response.headers['Content-Type'] = content_type
        response.headers["Content-Disposition"] = "attachment; filename={}".format(quote(filename))
        return response


def get_task_data(task_id):
    try:
        return task_repo.find_by_id(task_id)
    except Exception as e:
        import logging
        logging.getLogger(__name__).error(f"Failed to get task data for task_id={task_id}: {e}")


def get_ip_data(task_id):
    return ip_repo.find_by_task(task_id)


def get_site_data(task_id):
    return site_repo.find_by_task(task_id)


def get_domain_data(task_id):
    return domain_repo.find_by_task(task_id)


def port_service_product_statist(task_id):
    ip_data = get_ip_data(task_id)
    total = 0
    port_info_list = []
    for item in ip_data:
        if not item["port_info"]:
            continue
        port_info_list.extend(item["port_info"])
        total += len(item["port_info"])

    counter = Counter([info["port_id"] for info in port_info_list])
    top_20 = counter.most_common(20)
    port_percent_list = []
    for port_info in top_20:
        port_id, amount = port_info
        item = {
            "port_id" : port_id,
            "amount" : amount,
            "percent" : "{:.2f}%".format((amount *100.0 ) / total)
        }
        port_percent_list.append(item)

    service_name_list = []
    for info in port_info_list:
        if  not  info.get("product"):
            continue
        if info["product"] or info["version"]:
            service_name = info["service_name"]
            if service_name == "https-alt":
                service_name = "https"

            service_name_list.append(service_name)

    service_top_20 = Counter(service_name_list).most_common(20)
    service_percent_list = []
    for port_info in service_top_20:
        service_name, amount = port_info
        item = {
            "service_name" : service_name,
            "amount" : amount,
            "percent" : "{:.2f}%".format((amount *100.0 ) / len(service_name_list))
        }
        service_percent_list.append(item)



    product_name_list = []
    for info in port_info_list:
        if not info.get("product"):
            continue
        product = info["product"]
        if product and "**" not in product:
            product = product.strip()
            product_name_list.append(product)

    product_top_20 = Counter(product_name_list).most_common(20)
    product_percent_list = []
    for info in product_top_20:
        product, amount = info
        item = {
            "product" : product,
            "amount" : amount,
            "percent" : "{:.2f}%".format((amount *100.0 ) / len(product_name_list))
        }
        product_percent_list.append(item)

    statist = {
        "port_total": total, #端口开放总数
        "port_percent_list": port_percent_list, #端口开放 top 20比例详情
        "service_total": len(service_name_list),  #系统服务类别总数
        "service_percent_list": service_percent_list, #系统服务类别 top 20比例详情
        "product_total": len(product_name_list), #产品种类总数
        "product_percent_list": product_percent_list ##产品种类总数 top 20比例详情
    }
    return statist



class SaveTask(object):
    """docstring for ClassName"""

    def __init__(self, task_id):
        self.task_id = task_id
        self.wb = Workbook()
        self.is_ip_task = False

    def set_style(self, ws):
        font = Font(name="Consolas", color="111111")
        column = "ABCDEFGHIJKLMNO"
        for x in column:
            for y in range(1, 256):
                ws["{}{}".format(x,y)].font = font

    def build_service_xl(self):
        ws = self.wb.create_sheet(title="系统服务")
        ws.column_dimensions['A'].width = 22.0
        ws.column_dimensions['B'].width = 10.0
        ws.column_dimensions['C'].width = 20.0
        ws.column_dimensions['D'].width = 40.0

        column_tilte = ["IP", "端口","服务", "产品", "版本"]
        ws.append(column_tilte)
        for item in get_ip_data(self.task_id):
            for port_info in item["port_info"]:
                row = []
                row.append(item["ip"])
                row.append("{}".format(port_info["port_id"]))
                row.append(port_info["service_name"])
                row.append(port_info.get("product", ""))
                row.append(port_info.get("version", ""))
                ws.append(row)

        self.set_style(ws)

    def build_ip_xl(self):
        ws = self.wb.create_sheet(title="IP")
        ws.column_dimensions['A'].width = 22.0
        ws.column_dimensions['B'].width = 50.0
        ws.column_dimensions['C'].width = 10.0
        ws.column_dimensions['D'].width = 25.0
        ws.column_dimensions['E'].width = 55.0
        if self.is_ip_task:
            ws.column_dimensions['F'].width = 55.0
            column_tilte = ["IP", "端口信息", "开放端口数目", "geo", "as 编号", "操作系统"]
            ws.append(column_tilte)
            for item in get_ip_data(self.task_id):
                row = []
                row.append(item["ip"])

                port_ids = [str(x["port_id"]) for x in item["port_info"]]
                row.append(" \r\n".join(port_ids))
                row.append(len(item["port_info"]))
                if "country_name" in item["geo_city"]:
                    row.append("{}/{}".format(item["geo_city"]["country_name"],
                                              item["geo_city"]["region_name"]))
                    row.append(item["geo_asn"].get("organization", ""))
                else:
                    row.append("")
                    row.append("")

                osname = ""
                if item.get("os_info"):
                    osname = item["os_info"]["name"]
                row.append(osname)
                ws.append(row)
        else:
            ws.column_dimensions['F'].width = 60.0
            ws.column_dimensions['G'].width = 40.0
            ws.column_dimensions['H'].width = 40.0
            ws.column_dimensions['I'].width = 20.0
            column_tilte = ["IP", "端口信息", "开放端口数目", "geo", "as 编号"]
            column_tilte.append("domain")
            column_tilte.append("操作系统")
            column_tilte.append("CDN")
            column_tilte.append("类别")
            ws.append(column_tilte)
            for item in get_ip_data(self.task_id):
                row = []
                row.append(item["ip"])

                port_ids = [str(x["port_id"]) for x in item["port_info"]]
                row.append(" \r\n".join(port_ids))

                row.append(len(item["port_info"]))
                if "country_name" in item["geo_city"]:
                    row.append("{}/{}".format(item["geo_city"]["country_name"],
                                              item["geo_city"]["region_name"]))
                    row.append(item["geo_asn"].get("organization", ""))
                else:
                    row.append("")
                    row.append("")

                row.append(" \r\n".join(item.get("domain", [])))

                osname = ""
                if item.get("os_info"):
                    osname = item["os_info"]["name"]
                row.append(osname)
                row.append(item.get("cdn_name", ""))
                row.append(item.get("ip_type", ""))
                ws.append(row)

        self.set_style(ws)

    def ignore_illegal(self, content):
        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')
        content = ILLEGAL_CHARACTERS_RE.sub(r'', content)
        return content

    def build_site_xl(self):
        ws = self.wb.active
        ws.column_dimensions['A'].width = 35.0
        ws.column_dimensions['B'].width = 40.0
        ws.column_dimensions['C'].width = 60.0
        ws.column_dimensions['D'].width = 20.0
        ws.column_dimensions['E'].width = 30.0
        ws.title = "站点"
        column_tilte = ["site", "title", "指纹", "状态码", "favicon hash"]
        ws.append(column_tilte)
        for item in get_site_data(self.task_id):
            row = []
            row.append(self.ignore_illegal(item["site"]))
            row.append(self.ignore_illegal(item["title"]))
            row.append(" \r\n".join([self.ignore_illegal(x["name"]) for x in item["finger"]]))
            row.append(item["status"])
            row.append(item["favicon"].get("hash", ""))
            ws.append(row)

        self.set_style(ws)

    def build_domain_xl(self):
        ws = self.wb.create_sheet(title="域名")
        ws.column_dimensions['A'].width = 30.0
        ws.column_dimensions['B'].width = 20.0
        ws.column_dimensions['C'].width = 50.0
        ws.column_dimensions['D'].width = 50.0

        column_tilte = ["域名", "解析类型", "记录值","关联ip"]

        ws.append(column_tilte)
        for item in get_domain_data(self.task_id):
            row = []
            row.append(item["domain"])
            row.append(item["type"])
            row.append(" \r\n".join(item["record"]))
            row.append(" \r\n".join(item["ips"]))
            ws.append(row)

        self.set_style(ws)

    def build_statist(self):
        statist = port_service_product_statist(self.task_id)
        ws = self.wb.create_sheet(title="资产统计")
        ws.column_dimensions['A'].width = 20.0
        ws.column_dimensions['F'].width = 20.0
        ws.column_dimensions['K'].width = 40.0
        ws["A1"] = "端口信息统计"
        ws["F1"] = "系统服务信息统计"
        ws["K1"] = "软件产品信息统计"

        ports = ["端口", "数量", "占比"]
        port_percent_list = statist["port_percent_list"]
        port_total = statist["port_total"]
        for port_info in port_percent_list:
            ports.append(port_info["port_id"])
            ports.append(port_info["amount"])
            ports.append(port_info["percent"])

        cnt = 0
        for row in range(5, 27):
            for col in range(1, 4):
                if cnt >= len(ports):
                    continue
                ws.cell(column=col, row=row, value=ports[cnt])
                cnt += 1

        ws["A27"] = "端口开放总数"
        ws["A28"] = port_total

        services = ["系统服务", "数量", "占比"]
        service_percent_list = statist["service_percent_list"]
        if len(service_percent_list) >= 0:
            service_total = statist["service_total"]
            for port_info in service_percent_list:
                services.append(port_info["service_name"])
                services.append(port_info["amount"])
                services.append(port_info["percent"])
            cnt = 0
            for row in range(5, 27):
                for col in range(6, 9):
                    if cnt >= len(services):
                        continue
                    ws.cell(column=col, row=row, value=services[cnt])
                    cnt += 1
            ws["F27"] = "系统服务类别总数"
            ws["F28"] = service_total

        product = ["产品", "数量", "占比"]
        product_percent_list = statist["product_percent_list"]
        if len(product_percent_list) >= 0:
            product_total = statist["product_total"]
            for port_info in product_percent_list:
                product.append(port_info["product"])
                product.append(port_info["amount"])
                product.append(port_info["percent"])
            cnt = 0
            for row in range(5, 27):
                for col in range(11, 14):
                    if cnt >= len(product):
                        continue
                    ws.cell(column=col, row=row, value=product[cnt])
                    cnt += 1
            ws["K27"] = "产品类别总数"
            ws["K28"] = product_total

        self.set_style(ws)

    def run(self):
        task_data = get_task_data(self.task_id)
        if not task_data:
            logger.warning("Export task not found: %s", self.task_id)
            return

        domain = task_data["target"].replace("/", "_")[:20]

        if re.findall(r"\b\d+\.\d+\.\d+\.\d+", domain):
            self.is_ip_task = True
        else:
            if task_data.get("type", "") == "ip":
                self.is_ip_task = True

        self.build_site_xl()
        self.build_ip_xl()
        self.build_service_xl()
        if not self.is_ip_task:
            self.build_domain_xl()

        self.build_statist()

        return save_virtual_workbook(self.wb)


def export_arl(task_id):
    task_id = task_id.strip()
    save = SaveTask(task_id)
    return save.run()


# -- 导出格式写入器 --

def _write_xlsx(task_data, task_id):
    """XLSX 格式导出（保持兼容 export_arl）"""
    task_id = task_id.strip()
    save = SaveTask(task_id)
    data = save.run()
    return data, 'application/octet-stream'


def _write_json(task_data, task_id):
    """JSON 格式导出"""
    ip_list = list(get_ip_data(task_id))
    domain_list = list(get_domain_data(task_id))
    site_list = list(get_site_data(task_id))
    statist = port_service_product_statist(task_id)

    result = {
        'task': {
            'name': task_data.get('name', ''),
            'target': task_data.get('target', ''),
            'status': task_data.get('status', ''),
        },
        'statistics': statist,
        'ips': _serialize_list(ip_list),
        'domains': _serialize_list(domain_list),
        'sites': _serialize_list(site_list),
    }
    data = json.dumps(result, ensure_ascii=False, default=str, indent=2)
    return data.encode('utf-8'), 'application/json'


def _serialize_list(items):
    """将 MongoDB ObjectId 转换为字符串"""
    from bson import ObjectId
    result = []
    for item in items:
        serialized = {}
        for k, v in item.items():
            if isinstance(v, ObjectId):
                serialized[k] = str(v)
            elif isinstance(v, list):
                serialized[k] = [_serialize_value(x) for x in v]
            elif isinstance(v, dict):
                serialized[k] = {kk: _serialize_value(vv) for kk, vv in v.items()}
            else:
                serialized[k] = v
        result.append(serialized)
    return result


def _serialize_value(v):
    from bson import ObjectId
    if isinstance(v, ObjectId):
        return str(v)
    if isinstance(v, dict):
        return {kk: _serialize_value(vv) for kk, vv in v.items()}
    return v


# 注册导出格式（新格式通过 register_format('name', writer_func) 添加即可）
_EXPORT_FORMATS['xlsx'] = _write_xlsx
_EXPORT_FORMATS['json'] = _write_json
